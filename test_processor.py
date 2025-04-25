import pandas as pd
import numpy as np
from datetime import datetime
from pytz import timezone
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from io import BytesIO
import traceback
from logger import registrar_log

# Apenas os 3 status válidos que serão aceitos no processamento final
STATUS_VALIDOS_FINAIS = ['Passed', 'Not Executed', 'Failed']

# Status que serão normalizados (podem existir no arquivo original mas serão convertidos)
STATUS_NORMALIZAVEIS = [
    'Passed', 'Not Executed', 'Failed',
    'passed', 'not executed', 'failed',
    'PASSED', 'NOT EXECUTED', 'FAILED',
    'Not executed'  # Variação específica vista na imagem
]

# Status que serão explicitamente excluídos
STATUS_EXCLUIDOS = ['A validar', 'A Validar', 'Validar', 'validar', 'VALIDAR']

CORES_STATUS = {
    'Passed': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
    'Not Executed': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),
    'Failed': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
}

# Para o mapeamento de cores, criar aliases para diferentes formas de escrita
for status in ['passed', 'PASSED']:
    CORES_STATUS[status] = CORES_STATUS['Passed']
    
for status in ['not executed', 'NOT EXECUTED', 'Not executed']:
    CORES_STATUS[status] = CORES_STATUS['Not Executed']
    
for status in ['failed', 'FAILED']:
    CORES_STATUS[status] = CORES_STATUS['Failed']

# Colunas a serem ignoradas no processamento
COLUNAS_IGNORAR = ['ID Fluxo', 'Planejamento', 'Prioridade', 'Obervação']

# Mapeamento de colunas: chave = possíveis nomes na planilha, valor = nome padronizado
MAPEAMENTO_COLUNAS = {
    'Status': 'Status',
    'Status do Teste QD': 'Status',
    'Status de Teste': 'Status',
    'N° INC': 'N° INC',
    'Notas QD': 'Observação',
    'Obervação': 'Observação',
    'Observação': 'Observação'
}

COLUNAS_DESTINO = [
    'Data', 'Frente', 'Canal', 'Plataforma', 'Tipo de Plano', 'Plano',
    'Característica da massa', 'Entrypoint', 'Funcionalidade', 'Cenário',
    'Resultado esperado', 'Status', 'N° INC'
]

def processar_testes(arquivo_caderno, arquivo_diario, data_manual=None):
    """
    Processa e mescla os arquivos de teste no arquivo diário existente.
    
    Args:
        arquivo_caderno: Arquivo Excel contendo os testes nas abas "Caderno App Vivo" e "Caderno Web B2C"
        arquivo_diario: Arquivo Excel de acompanhamento diário com a aba "B2C"
        data_manual: Data no formato DD/MM/YYYY para os registros (opcional)
        
    Returns:
        tuple: (BytesIO do arquivo processado, quantidade de registros adicionados)
        
    Raises:
        Exception: Erro durante o processamento dos arquivos
    """
    try:
        registrar_log("Iniciando processamento de arquivos de teste", "info")
        all_data = []  # Usar uma lista para armazenar todos os registros
        
        # Processar cada aba do caderno de testes
        for sheet_name in ['Caderno App Vivo', 'Caderno Web B2C']:
            try:
                registrar_log(f"Processando aba {sheet_name}", "info")
                try:
                    # Tentar ler a aba específica
                    df = pd.read_excel(
                        arquivo_caderno,
                        sheet_name=sheet_name,
                        engine='openpyxl',
                        dtype=str
                    )
                except Exception as sheet_error:
                    registrar_log(f"Erro ao ler aba {sheet_name}: {str(sheet_error)}", "erro")
                    # Tentar identificar todas as abas disponíveis
                    xls = pd.ExcelFile(arquivo_caderno)
                    available_sheets = xls.sheet_names
                    registrar_log(f"Abas disponíveis no arquivo: {', '.join(available_sheets)}", "info")
                    
                    # Se não conseguir encontrar a aba específica, tente usar a primeira aba
                    if len(available_sheets) > 0:
                        registrar_log(f"Tentando usar a primeira aba disponível: {available_sheets[0]}", "info")
                        df = pd.read_excel(
                            arquivo_caderno,
                            sheet_name=available_sheets[0],
                            engine='openpyxl',
                            dtype=str
                        )
                    else:
                        raise Exception(f"Não foi possível encontrar nenhuma aba válida no arquivo!")
                
                # Registrar as colunas encontradas para diagnóstico
                registrar_log(f"Colunas encontradas na aba {sheet_name}: {', '.join(df.columns.tolist())}", "info")
                
                # Verificar se há alguma coluna de status
                coluna_status = None
                for col in df.columns:
                    if col in MAPEAMENTO_COLUNAS and MAPEAMENTO_COLUNAS[col] == 'Status':
                        coluna_status = col
                        registrar_log(f"Coluna de status encontrada: '{col}'", "info")
                        break
                
                if not coluna_status:
                    registrar_log(f"Nenhuma coluna de status reconhecida na aba {sheet_name}", "aviso")
                    continue
                
                # Renomear as colunas com base no mapeamento
                colunas_renomeadas = {}
                for col in df.columns:
                    if col in MAPEAMENTO_COLUNAS:
                        colunas_renomeadas[col] = MAPEAMENTO_COLUNAS[col]
                
                # Aplicar renomeação e remover colunas ignoradas
                df = df.rename(columns=colunas_renomeadas)
                for col in COLUNAS_IGNORAR:
                    if col in df.columns:
                        df = df.drop(columns=[col])
                
                # Verificar e tratar valores nulos no status
                if 'Status' in df.columns:
                    # Preencher valores nulos e converter para string
                    df['Status'] = df['Status'].fillna('').astype(str)
                    registrar_log(f"Valores nulos na coluna Status tratados", "info")
                
                # Resolver problema com índices duplicados: converter para dicionário e depois para lista
                df_records = df.to_dict('records')
                all_data.extend(df_records)
                
                registrar_log(f"Aba {sheet_name} processada: {len(df_records)} linhas", "info")
                
            except Exception as e:
                registrar_log(f"Erro ao processar aba {sheet_name}: {str(e)}", "erro")
                registrar_log(f"Detalhes: {traceback.format_exc()}", "erro")
                continue
        
        # Verificar se há dados para processar
        if not all_data:
            msg = "Nenhuma aba válida encontrada no caderno de testes!"
            registrar_log(msg, "erro")
            raise Exception(msg)
            
        # Criar um novo DataFrame a partir dos registros combinados
        df_combined = pd.DataFrame(all_data)
        registrar_log(f"Total de linhas combinadas: {len(df_combined)}", "info")
        
        # Verificar se a coluna Status existe
        if 'Status' not in df_combined.columns:
            msg = "Coluna 'Status' não encontrada nos dados. Verifique se o arquivo tem as colunas corretas."
            registrar_log(msg, "erro")
            raise Exception(msg)
        
        # Tratamento de valores nulos/vazios na coluna Status
        registrar_log(f"Verificando valores nulos na coluna Status", "info")
        valores_nulos = df_combined['Status'].isna().sum()
        if valores_nulos > 0:
            registrar_log(f"Encontrados {valores_nulos} valores nulos na coluna Status", "aviso")
            df_combined['Status'] = df_combined['Status'].fillna('')
            
        # Verificar tipos de dados na coluna Status
        tipos_status = df_combined['Status'].apply(type).unique()
        registrar_log(f"Tipos de dados na coluna Status: {[t.__name__ for t in tipos_status]}", "info")
        
        # Garantir que Status seja string antes de processar
        df_combined['Status'] = df_combined['Status'].astype(str)
        
        # Padronizar status (remover espaços extras e normalizar case)
        df_combined['Status'] = df_combined['Status'].str.strip().str.title()
        
        # Registrar valores únicos de status para diagnóstico
        status_valores = df_combined['Status'].unique().tolist()
        registrar_log(f"Valores únicos de status encontrados: {', '.join([str(s) for s in status_valores])}", "info")
        
        # Normalizar valores específicos (caso 'not executed' esteja escrito diferente) de forma segura
        try:
            # Normalizar valores para os três status válidos finais
            # Primeiro, normaliza 'Not Executed'
            mask = df_combined['Status'].str.lower().str.contains('not executed', na=False)
            df_combined.loc[mask, 'Status'] = 'Not Executed'
            
            # Normaliza 'Passed'
            mask = df_combined['Status'].str.lower() == 'passed'
            df_combined.loc[mask, 'Status'] = 'Passed'
            
            # Normaliza 'Failed'
            mask = df_combined['Status'].str.lower() == 'failed'
            df_combined.loc[mask, 'Status'] = 'Failed'
            
            # Registro para debug
            status_apos_normalizacao = df_combined['Status'].unique().tolist()
            registrar_log(f"Status após normalização: {', '.join([str(s) for s in status_apos_normalizacao])}", "info")
            
            # Verificar presença de status indesejados
            status_indesejados = [s for s in status_apos_normalizacao if s not in STATUS_VALIDOS_FINAIS]
            if status_indesejados:
                registrar_log(f"Status indesejados encontrados e que serão excluídos: {', '.join(status_indesejados)}", "aviso")
                
            # Verificar explicitamente se temos status "A validar" ou variações
            status_a_validar = [s for s in status_apos_normalizacao if any(excluido.lower() in s.lower() for excluido in STATUS_EXCLUIDOS)]
            if status_a_validar:
                registrar_log(f"Status 'A validar' encontrados e que serão excluídos: {', '.join(status_a_validar)}", "aviso")
        except Exception as e:
            registrar_log(f"Erro ao normalizar valores de status: {str(e)}", "erro")
            # Continuar o processamento mesmo que a normalização falhe
        
        # Filtrar APENAS pelos três status válidos finais com checagem estrita
        registrar_log(f"Filtrando apenas pelos status: {', '.join(STATUS_VALIDOS_FINAIS)}", "info")
        
        # Antes de filtrar, garantir que nenhum status excluído seja erroneamente incluído
        for status_excluido in STATUS_EXCLUIDOS:
            # Remover explicitamente qualquer registro com status a ser excluído
            mask_excluir = df_combined['Status'].str.contains(status_excluido, case=False, na=False)
            if mask_excluir.any():
                qtd_excluidos = mask_excluir.sum()
                registrar_log(f"Excluindo {qtd_excluidos} registros com status '{status_excluido}'", "info")
                df_combined = df_combined[~mask_excluir].copy()
        
        # Agora aplicar o filtro de status válidos
        df_filtrado = df_combined[df_combined['Status'].isin(STATUS_VALIDOS_FINAIS)].copy()
        registrar_log(f"Linhas com status válido: {len(df_filtrado)} de {len(df_combined)} total", "info")
        
        # Verificação final para garantir que não temos status indesejados
        status_finais = df_filtrado['Status'].unique().tolist()
        registrar_log(f"Status após filtragem final: {', '.join([str(s) for s in status_finais])}", "info")
        
        if df_filtrado.empty:
            msg = f"Nenhum teste com status válido encontrado para processar! Aceitos apenas: {', '.join(STATUS_VALIDOS_FINAIS)}"
            registrar_log(msg, "aviso")
            
            # Mostrar os status encontrados para ajudar no diagnóstico
            status_valores = df_combined['Status'].unique().tolist()
            if status_valores:
                msg += f" Status encontrados: {', '.join([str(s) for s in status_valores])}"
            
            raise Exception(msg)
            
        # Verificar de novo se há algum status inválido
        for status in df_filtrado['Status'].unique():
            if status not in STATUS_VALIDOS_FINAIS:
                registrar_log(f"ALERTA: Status inválido '{status}' ainda presente após filtragem!", "erro")
                # Remover este status específico
                df_filtrado = df_filtrado[df_filtrado['Status'] != status]
        
        # Adicionar data aos registros
        data = data_manual if data_manual else datetime.now(timezone('America/Sao_Paulo')).strftime('%d/%m/%Y')
        df_filtrado.insert(0, 'Data', data)
        registrar_log(f"Data adicionada aos registros: {data}", "info")
        
        # Carregar o arquivo diário
        try:
            registrar_log("Carregando arquivo diário", "info")
            wb = load_workbook(BytesIO(arquivo_diario.read()))
            
            # Verificar se a aba B2C existe
            if 'B2C' not in wb.sheetnames:
                msg = "Aba 'B2C' não encontrada no arquivo diário!"
                registrar_log(msg, "erro")
                raise Exception(msg)
                
            ws = wb['B2C']
            registrar_log("Aba B2C encontrada e carregada", "info")
        except Exception as e:
            registrar_log(f"Erro ao carregar arquivo diário: {str(e)}", "erro")
            raise Exception(f"Erro ao carregar arquivo diário: {str(e)}")
        
        # Encontrar a última linha com dados
        ultima_linha = ws.max_row
        while ultima_linha > 0 and ws.cell(row=ultima_linha, column=1).value is None:
            ultima_linha -= 1
            
        registrar_log(f"Última linha com dados: {ultima_linha}", "info")
        
        # Obter cabeçalho do arquivo diário
        header = [cell.value for cell in ws[1]]
        registrar_log(f"Cabeçalho obtido: {len(header)} colunas", "info")
        
        # Mapear as colunas do DataFrame para corresponder ao cabeçalho do arquivo
        try:
            df_mapped = df_filtrado.reindex(columns=header, fill_value='')
        except Exception as e:
            registrar_log(f"Erro ao mapear colunas: {str(e)}", "erro")
            # Tentar corrigir problemas de índice
            registrar_log("Tentando abordagem alternativa para mapear colunas", "info")
            df_mapped = pd.DataFrame(columns=header)
            for col in header:
                if col in df_filtrado.columns:
                    df_mapped[col] = df_filtrado[col]
        
        # Adicionar os novos registros ao arquivo
        registrar_log(f"Adicionando {len(df_mapped)} novos registros", "info")
        for r_idx, row in enumerate(dataframe_to_rows(df_mapped, index=False, header=False), 1):
            nova_linha = ultima_linha + r_idx
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=nova_linha, column=c_idx, value=value)
                
                # Aplicar formatação de cor para o status
                if c_idx <= len(header) and header[c_idx-1] == 'Status':
                    status = str(value).strip().title() if value is not None else ''
                    cell.fill = CORES_STATUS.get(status, PatternFill())
        
        # Salvar o resultado
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        registrar_log(f"Processamento concluído com sucesso: {len(df_mapped)} registros adicionados", "info")
        return output, len(df_mapped)

    except Exception as e:
        registrar_log(f"Erro crítico no processamento de testes: {str(e)}", "erro")
        registrar_log(f"Detalhes: {traceback.format_exc()}", "erro")
        raise Exception(f"Erro ao processar testes: {str(e)}") 