# -*- coding: utf-8 -*-
import pandas as pd
import streamlit as st
import os
import json
import traceback
from logger import configurar_logs, registrar_log
from datetime import datetime, timedelta, time
from pytz import timezone
from chg_comparator import comparar_chgs
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl import load_workbook
from io import BytesIO
from copy import deepcopy
from openpyxl.formatting.rule import Rule
from PIL import Image
# Removendo a importação de gera_relatorio para evitar conflitos
# from gera_relatorio import gerar_relatorio, processar_json
# Importando o novo módulo para a página de relatório de incidentes
from incident_report_page import render_incident_report_page
# Importando o novo módulo para a página de processamento de testes
from test_processor_page import render_test_processor_page

# Carrega o favicon
favicon = Image.open("spread_logo.png")

# Configurações iniciais e estilo
st.set_page_config(
    page_title="QD Apps - Sustentação",
    page_icon=favicon,
    layout="wide",
    initial_sidebar_state="collapsed"
)

# Estilo CSS personalizado
st.markdown("""
    <style>
    .main {
        padding: 0rem 0.5rem;
    }
    .stTabs [data-baseweb="tab-list"] {
        gap: 1px;
    }
    .stTabs [data-baseweb="tab"] {
        padding: 8px 16px;
        background-color: #f0f2f6;
        border-radius: 4px 4px 0 0;
        font-size: 0.9em;
    }
    .stTabs [aria-selected="true"] {
        background-color: #1f61d9;
        color: white;
    }
    .uploadedFile {
        border: 1px solid #ccc;
        border-radius: 5px;
        padding: 15px;
        margin: 8px 0;
    }
    .stButton>button {
        background-color: #1f61d9;
        color: white;
        border-radius: 5px;
        padding: 8px 20px;
        font-weight: 500;
        border: none;
        transition: all 0.3s ease;
        font-size: 0.9em;
    }
    .stButton>button:hover {
        background-color: #1a4fa8;
        box-shadow: 0 2px 4px rgba(0,0,0,0.1);
    }
    .success-message {
        padding: 15px;
        border-radius: 8px;
        background-color: #e7f3e7;
        border-left: 4px solid #28a745;
        font-size: 0.9em;
    }
    .warning-message {
        padding: 15px;
        border-radius: 8px;
        background-color: #fff3cd;
        border-left: 4px solid #ffc107;
        font-size: 0.9em;
    }
    .error-message {
        padding: 15px;
        border-radius: 8px;
        background-color: #f8d7da;
        border-left: 4px solid #dc3545;
        font-size: 0.9em;
    }
    /* Reduzir espaçamento geral */
    .st-emotion-cache-1y4p8pa {
        padding: 1rem 0.5rem;
    }
    /* Reduzir tamanho dos headers */
    h1 {
        font-size: 1.8em !important;
    }
    h2 {
        font-size: 1.5em !important;
        margin-bottom: 15px !important;
    }
    h3 {
        font-size: 1.2em !important;
        margin-bottom: 10px !important;
    }
    /* Reduzir padding dos containers */
    .st-emotion-cache-1r6slb0 {
        padding: 10px !important;
    }
    /* Ajustar tamanho do logo e cabeçalho */
    .st-emotion-cache-1v0mbdj {
        margin-bottom: 15px !important;
    }
    </style>
""", unsafe_allow_html=True)

# Logo e título com estilo moderno
st.markdown(
    f"""
    <div style="text-align: center; padding: 15px 0; background: linear-gradient(90deg, #1f61d9, #1a4fa8); border-radius: 8px; margin-bottom: 20px;">
        <img src="https://spread.com.br/wp-content/uploads/2023/10/logo-white.svg" style="max-width:200px; margin-bottom:15px">
        <h1 style="color: white; font-size: 2em; font-weight: 600;">QD Apps - Sustentação</h1>
    </div>
    """, 
    unsafe_allow_html=True
)

configurar_logs()

# ========== Funções Principais ==========
def map_status_emoji(status):
    emoji_map = {
        'Novo': '🆕', 'Agendado': '🕔', 'Implementar': '💻',
        'Em Execução': '⚙️', 'Revisão': '⚠️', 'Cancelada': '❌',
        'Finalizada': '✅', 'CHG com Indisponibilidade': '📵', 'Avaliar': '⚠️'
    }
    return emoji_map.get(status, status)

def processar_dados(uploaded_file):
    try:
        registrar_log("Iniciando processamento do arquivo", "info")
        
        # Configurar timezone de Brasília
        tz_brasilia = timezone('America/Sao_Paulo')
        
        # Obter data e hora atual
        agora = datetime.now(tz_brasilia)
        hoje = agora.date()
        amanha = hoje + timedelta(days=1)
        
        # Criar strings de datas para comparação
        hoje_str = hoje.strftime('%Y-%m-%d')
        amanha_str = amanha.strftime('%Y-%m-%d')
        
        registrar_log(f"Data de hoje: {hoje_str}, Data de amanhã: {amanha_str}", "info")
        
        try:
            df1 = pd.read_excel(uploaded_file, sheet_name='CHGs', engine='openpyxl')
            registrar_log(f"Leitura da aba CHGs concluída: {len(df1)} linhas", "info")
        except Exception as e:
            registrar_log(f"Erro na leitura da aba CHGs: {str(e)}", "erro")
            df1 = pd.DataFrame()
            
        try:
            df2 = pd.read_excel(uploaded_file, sheet_name='CHGs II', engine='openpyxl')
            registrar_log(f"Leitura da aba CHGs II concluída: {len(df2)} linhas", "info")
        except Exception as e:
            registrar_log(f"Erro na leitura da aba CHGs II: {str(e)}", "erro")
            df2 = pd.DataFrame()
        
        if df1.empty and df2.empty:
            registrar_log("Ambas abas estão vazias ou não foram lidas corretamente", "erro")
            st.error("Não foi possível ler dados do arquivo. Verifique se o formato está correto.")
            return pd.DataFrame()
            
        df = pd.concat([df1, df2], ignore_index=True)
        registrar_log(f"Total de linhas após concatenação: {len(df)}", "info")
        
        colunas = ['Número', 'Descrição resumida', 'Status', 'Tipo de Indisponibilidade',
                 'Data de início planejada', 'Data de término planejada', 'IC Impactado', 
                 'Grupo de atribuição', 'Observação (Time Mudanças)', 'Enviar Keep']
        
        # Verifica se todas as colunas existem
        colunas_faltantes = [col for col in colunas if col not in df.columns]
        if colunas_faltantes:
            msg_erro = f"Colunas faltantes no arquivo: {', '.join(colunas_faltantes)}"
            registrar_log(msg_erro, "erro")
            st.error(msg_erro)
            return pd.DataFrame()
        
        df = df[colunas].copy()
        
        # Registrar informações sobre os tipos de dados na coluna 'Data de início planejada'
        registrar_log(f"Tipo de dados na coluna 'Data de início planejada': {df['Data de início planejada'].dtype}", "info")
        
        # Converter datas para strings para evitar problemas de conversão
        try:
            # Verifica se a coluna já contém strings
            if pd.api.types.is_string_dtype(df['Data de início planejada']):
                registrar_log("A coluna 'Data de início planejada' já contém strings", "info")
                # Converter string para datetime primeiro
                df['Data de início planejada'] = pd.to_datetime(df['Data de início planejada'], errors='coerce')
                registrar_log("Conversão de strings para datetime concluída", "info")
            else:
                # Tenta converter para datetime e depois para string
                df['Data de início planejada'] = pd.to_datetime(df['Data de início planejada'], errors='coerce')
                registrar_log("Conversão de 'Data de início planejada' para datetime concluída", "info")
            
            # Verificar se há valores nulos após a conversão
            if df['Data de início planejada'].isna().any():
                num_nulos = df['Data de início planejada'].isna().sum()
                registrar_log(f"Atenção: {num_nulos} valores não puderam ser convertidos para data", "aviso")
                # Remover linhas com datas nulas para evitar problemas
                df = df.dropna(subset=['Data de início planejada'])
                registrar_log(f"Linhas com datas nulas removidas. Restantes: {len(df)}", "info")
            
            # Verificar se ainda existem linhas após a filtragem
            if df.empty:
                registrar_log("Todas as linhas foram removidas durante a limpeza de datas", "erro")
                st.error("Não foi possível processar o arquivo: todas as datas são inválidas.")
                return pd.DataFrame()
            
            # Guarda a coluna original para exibição
            df['Data de início original'] = df['Data de início planejada'].copy()
            
            # Cria uma coluna só com a data em formato string (YYYY-MM-DD)
            df['data_inicio_str'] = df['Data de início planejada'].dt.strftime('%Y-%m-%d')
            registrar_log(f"Criação da coluna 'data_inicio_str' concluída", "info")
            
            # Cria uma coluna só com a hora em formato numérico (24h)
            df['hora_inicio'] = df['Data de início planejada'].dt.hour
            registrar_log(f"Criação da coluna 'hora_inicio' concluída", "info")
            
            # Registrar amostra de algumas linhas para debug
            amostra = df[['Data de início planejada', 'data_inicio_str', 'hora_inicio']].head(3)
            registrar_log(f"Amostra de dados após conversão: {amostra.to_dict()}", "info")
            
            # Converter coluna Data de término planejada
            df['Data de término planejada'] = pd.to_datetime(df['Data de término planejada'], errors='coerce')
            
            # Verificar se há valores nulos após a conversão da data de término
            if df['Data de término planejada'].isna().any():
                num_nulos = df['Data de término planejada'].isna().sum()
                registrar_log(f"Atenção: {num_nulos} valores de data de término não puderam ser convertidos", "aviso")
                # Remover linhas com datas de término nulas
                df = df.dropna(subset=['Data de término planejada'])
                registrar_log(f"Linhas com datas de término nulas removidas. Restantes: {len(df)}", "info")
                
            # Verificar se ainda existem linhas após a filtragem
            if df.empty:
                registrar_log("Todas as linhas foram removidas durante a limpeza de datas de término", "erro")
                st.error("Não foi possível processar o arquivo: todas as datas de término são inválidas.")
                return pd.DataFrame()
        except Exception as e:
            msg_erro = f"Erro na conversão de datas: {str(e)}"
            registrar_log(msg_erro, "erro")
            registrar_log(f"Detalhes do erro: {traceback.format_exc()}", "erro")
            st.error(msg_erro)
            return pd.DataFrame()
        
        df['Observação (Time Mudanças)'] = df['Observação (Time Mudanças)'].fillna('')
        
        # Lógica de filtragem baseada em strings e valores numéricos
        try:
            # Filtro para hoje: data == hoje E hora >= 17
            hoje_filtro = (df['data_inicio_str'] == hoje_str) & (df['hora_inicio'] >= 17)
            registrar_log(f"Filtro para hoje criado: {hoje_filtro.sum()} linhas", "info")
            
            # Filtro para amanhã: data == amanhã E hora < 4
            amanha_filtro = (df['data_inicio_str'] == amanha_str) & (df['hora_inicio'] < 4)
            registrar_log(f"Filtro para amanhã criado: {amanha_filtro.sum()} linhas", "info")
            
            # Filtro para Enviar Keep
            if 'Enviar Keep' in df.columns:
                df['Enviar Keep'] = df['Enviar Keep'].astype(str)
                keep_filtro = df['Enviar Keep'].str.strip().str.lower() == 'sim'
                registrar_log(f"Filtro para 'Enviar Keep' criado: {keep_filtro.sum()} linhas", "info")
            else:
                registrar_log("Coluna 'Enviar Keep' não encontrada, considerando todas as linhas", "aviso")
                keep_filtro = pd.Series([True] * len(df))
            
            # Filtragem final
            df_filtrado = df[
                (hoje_filtro | amanha_filtro) &
                keep_filtro
            ]
            
            # Remover colunas auxiliares que não serão mostradas no relatório
            if 'data_inicio_str' in df_filtrado.columns:
                df_filtrado = df_filtrado.drop(columns=['data_inicio_str'])
            if 'hora_inicio' in df_filtrado.columns:
                df_filtrado = df_filtrado.drop(columns=['hora_inicio'])
            
            registrar_log(f"CHGs encontradas (hoje a partir das 17:00 e amanhã até 04:00): {len(df_filtrado)}", "info")
            return df_filtrado
        except Exception as e:
            msg_erro = f"Erro na filtragem de dados: {str(e)}"
            registrar_log(msg_erro, "erro")
            registrar_log(f"Detalhes do erro: {traceback.format_exc()}", "erro")
            st.error(msg_erro)
            return pd.DataFrame()
        
    except Exception as e:
        erro_detalhado = traceback.format_exc()
        st.error(f"Erro crítico: {str(e)}")
        registrar_log(f"Erro no processamento: {str(e)}", "erro")
        registrar_log(f"Detalhes do erro: {erro_detalhado}", "erro")
        return pd.DataFrame()

def gerar_relatorio(df):
    if df.empty:
        return "Nenhuma CHG encontrada para o dia de hoje com os filtros aplicados."
    
    relatorio = """💻 *REPORT STATUS CHGs – QD APPs* 💻  

Segue CHGs que serão executadas: 

"""
    
    for _, row in df.iterrows():
        tipo_indisponibilidade = str(row['Tipo de Indisponibilidade']).lower()
        indisponibilidade = "📵 " if "indisponibilidade parcial" in tipo_indisponibilidade or "indisponibilidade total" in tipo_indisponibilidade else "👍 "
        
        # Formatar datas com tratamento de erro
        try:
            data_inicio = row['Data de início planejada'].strftime('%d/%m/%Y %H:%M')
        except:
            data_inicio = "[Data inválida]"
            
        try:
            data_termino = row['Data de término planejada'].strftime('%d/%m/%Y %H:%M')
        except:
            data_termino = "[Data inválida]"
        
        relatorio += f"""*Mudança:* {row['Número']}
*✏ Descrição:* {row['Descrição resumida']}
*Tipo de Indisponibilidade:* {indisponibilidade}{row['Tipo de Indisponibilidade']}
*IC Impactado:* {row['IC Impactado']}
*Grupo de atribuição:* {row['Grupo de atribuição']}
*Início:* {data_inicio}
*Término:* {data_termino}
*Observação:* {row['Observação (Time Mudanças)']}\n\n"""

    relatorio += """*Legenda:*
⚠️ Ponto de Atenção
📵 CHG com Indisponibilidade
👍 Sem Indisponibilidade 


 QD Spread"""
    
    return relatorio

COLUNAS_ALVO = [
    'Plataforma', 'Tipo de Plano', 'Plano', 'Característica da massa',
    'Entrypoint', 'Funcionalidade', 'Cenário', 'Resultado esperado',
    'Status', 'N° INC'
]

# Removendo o código duplicado pois agora estamos usando o módulo test_processor.py
# As constantes abaixo são definidas no test_processor.py

# STATUS_VALIDOS = ['Passed', 'Not Executed', 'Failed']
# CORES_STATUS = {
#     'Passed': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
#     'Not Executed': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),
#     'Failed': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
# }

# COLUNAS_IGNORAR = ['ID Fluxo', 'Planejamento', 'Prioridade', 'Obervação']
# COLUNAS_DESTINO = [
#     'Data', 'Frente', 'Canal', 'Plataforma', 'Tipo de Plano', 'Plano',
#     'Característica da massa', 'Entrypoint', 'Funcionalidade', 'Cenário',
#     'Resultado esperado', 'Status', 'N° INC'
# ]

# def processar_testes(arquivo_caderno, arquivo_diario, data_manual=None):
#     """Processa e mescla os arquivos de teste no arquivo diário existente"""
#     try:
#         dfs = []
#         for sheet_name in ['Full Web', 'Priorizado']:
#             try:
#                 df = pd.read_excel(
#                     arquivo_caderno,
#                     sheet_name=sheet_name,
#                     engine='openpyxl',
#                     dtype=str
#                 )
#                 
#                 df = df.rename(columns={
#                     'Obervação': 'Observação',
#                     'Status': 'Status',
#                     'N° INC': 'N° INC'
#                 }).drop(columns=COLUNAS_IGNORAR, errors='ignore')
#                 
#                 dfs.append(df)
#                 
#             except Exception as e:
#                 st.warning(f"Erro ao processar aba {sheet_name}: {str(e)}")
#                 continue
#         
#         if not dfs:
#             st.error("Nenhuma aba válida encontrada!")
#             return None, 0
#             
#         df_combined = pd.concat(dfs, axis=0, ignore_index=True, sort=False)
#         df_combined['Status'] = df_combined['Status'].str.strip().str.title()
#         df_filtrado = df_combined[df_combined['Status'].isin(STATUS_VALIDOS)].copy()
#         
#         if df_filtrado.empty:
#             st.warning("Nenhum teste válido encontrado para processar!")
#             return None, 0
#         
#         data = data_manual if data_manual else datetime.now(timezone('America/Sao_Paulo')).strftime('%d/%m/%Y')
#         df_filtrado.insert(0, 'Data', data)
#         
#         wb = load_workbook(BytesIO(arquivo_diario.read()))
#         ws = wb['B2C']
#         
#         ultima_linha = ws.max_row
#         while ws.cell(row=ultima_linha, column=1).value is None:
#             ultima_linha -= 1
#         
#         header = [cell.value for cell in ws[1]]
#         df_mapped = df_filtrado.reindex(columns=header, fill_value='')
#         
#         for r_idx, row in enumerate(dataframe_to_rows(df_mapped, index=False, header=False), 1):
#             nova_linha = ultima_linha + r_idx
#             for c_idx, value in enumerate(row, 1):
#                 cell = ws.cell(row=nova_linha, column=c_idx, value=value)
#                 
#                 if header[c_idx-1] == 'Status':
#                     status = str(value).strip().title()
#                     cell.fill = CORES_STATUS.get(status, PatternFill())
#         
#         output = BytesIO()
#         wb.save(output)
#         output.seek(0)
#         
#         return output, len(df_filtrado)
# 
#     except Exception as e:
#         st.error(f"Erro crítico: {str(e)}")
#         raise

COLUNAS_OCORRENCIAS = [
    'Número', 'Incidentes secundários', 'Aberto', 'Prioridade', 'Estado',
    'Descrição resumida', 'Descrição', 'Aberto por', 'Atribuído a',
    'Canal impactado', 'IC Impactado', 'IC Causador', 'Problema', 'Status',
    'Sub Status', 'Código de resolução', 'Causa Origem', 'Causa provável',
    'Causado pela mudança', 'Anotações de resolução', 'Resolvido', 'Encerrado',
    'u_rpt_tempo_total_de_impacto'
]

# ========== NOVA FUNÇÃO ==========
def atualizar_ocorrencias(planilha_base, planilha_funcionais, planilha_criticos):
    """Atualiza a planilha de ocorrências com os dados das extrações, mantendo a formatação original.
    A aba "Funcionais" da planilha base será atualizada com os dados da aba "extração funcionais" do arquivo de extração funcionais,
    e a aba "Criticos NOW" será atualizada com os dados da aba "extração criticos" do arquivo de extração criticos.
    """
    try:
        wb_base = load_workbook(BytesIO(planilha_base.read()))
        wb_funcionais = load_workbook(BytesIO(planilha_funcionais.read()), data_only=True)
        wb_criticos = load_workbook(BytesIO(planilha_criticos.read()), data_only=True)

        def update_sheet(ws, new_data):
            # Salva o número original de linhas formatadas (considerando que a 1ª linha é o cabeçalho)
            old_max = ws.max_row
            # Limpa apenas os valores dos dados, mantendo a formatação
            for row in ws.iter_rows(min_row=2, max_row=old_max):
                for cell in row:
                    cell.value = None

            num_new_rows = len(new_data)
            # Atualiza os valores nas linhas já existentes
            for i, row_data in enumerate(new_data, start=2):
                for j, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=i, column=j)
                    cell.value = value

            # Se houver mais linhas novas que as formatadas, copia a formatação da linha anterior para as novas linhas
            if num_new_rows > (old_max - 1):
                max_col = ws.max_column
                for i in range(old_max+1, num_new_rows+2):
                    for j in range(1, max_col+1):
                        template = ws.cell(row=i-1, column=j)
                        new_cell = ws.cell(row=i, column=j)
                        if template.font:
                            new_cell.font = template.font.copy()
                        if template.border:
                            new_cell.border = template.border.copy()
                        if template.fill:
                            new_cell.fill = template.fill.copy()
                        new_cell.number_format = template.number_format
                    row_index = i - 2
                    if row_index < num_new_rows:
                        for j, value in enumerate(new_data[row_index], start=1):
                            cell = ws.cell(row=i, column=j)
                            cell.value = value

        # Atualiza a aba 'Funcionais'
        ws_base_funcionais = wb_base["Funcionais"]
        ws_extr_funcionais = wb_funcionais["extração funcionais"]
        dados_funcionais = [row for row in ws_extr_funcionais.iter_rows(min_row=2, values_only=True)]
        update_sheet(ws_base_funcionais, dados_funcionais)

        # Atualiza a aba 'Criticos NOW'
        ws_base_criticos = wb_base["Criticos NOW"]
        sheet_name = "extração criticos" if "extração criticos" in wb_criticos.sheetnames else "extração críticos"
        ws_extr_criticos = wb_criticos[sheet_name]
        dados_criticos = [row for row in ws_extr_criticos.iter_rows(min_row=2, values_only=True)]
        update_sheet(ws_base_criticos, dados_criticos)

        output = BytesIO()
        wb_base.save(output)
        output.seek(0)
        total_registros = len(dados_funcionais) + len(dados_criticos)
        return output, total_registros
    except Exception as e:
        st.error(f"Erro crítico: {str(e)}")
        raise


# ========== Interface Streamlit ==========
tabs = st.tabs([
    "📤 Gerador de Keep CHGs",
    "📊 Relatório de Incidentes",
    "📋 Processador de Testes",
    "⚙️ Sobre"
])

with tabs[0]:
    st.markdown("""
        <div style="background-color: #f8f9fa; padding: 20px; border-radius: 10px; margin-bottom: 20px;">
            <h2 style="color: #1f61d9; margin-bottom: 20px;">📤 Gerador de Keep CHGs</h2>
        </div>
    """, unsafe_allow_html=True)
    
    with st.container():
        uploaded_file = st.file_uploader(
            "Arraste ou clique para carregar o arquivo XLSX",
            type=["xlsx"],
            key="file_uploader",
            help="Selecione o arquivo Excel contendo as CHGs"
        )

        if 'ultimo_arquivo' in st.session_state and not uploaded_file:
            del st.session_state.ultimo_arquivo
            st.rerun()

        if uploaded_file:
            with st.spinner('Processando arquivo...'):
                df = processar_dados(uploaded_file)
                
                if not df.empty:
                    relatorio = gerar_relatorio(df)
                    st.markdown(f"""
                        <div class="success-message">
                            ✅ {len(df)} CHGs de hoje/amanhã processadas com sucesso!
                        </div>
                    """, unsafe_allow_html=True)
                    
                    st.text_area(
                        "Prévia do Relatório",
                        relatorio,
                        height=500,
                        help="Visualize o relatório antes de baixar"
                    )
                    
                    col1, col2, col3 = st.columns([1,2,1])
                    with col2:
                        st.download_button(
                            "⬇️ Baixar Relatório",
                            relatorio,
                            "CHGs_Report.txt",
                            use_container_width=True
                        )
                else:
                    st.markdown("""
                        <div class="warning-message">
                            ⚠️ Nenhuma CHG encontrada para hoje!
                        </div>
                    """, unsafe_allow_html=True)

# Nova aba para Relatório de Incidentes - usando o módulo separado
with tabs[1]:
    render_incident_report_page()

# Nova aba para Processador de Testes - usando o módulo separado
with tabs[2]:
    render_test_processor_page()

# A aba "Sobre" agora será a quarta aba
with tabs[3]:
    st.markdown("""
        <div style="background-color: #f8f9fa; padding: 20px; border-radius: 10px; margin-bottom: 20px;">
            <h2 style="color: #1f61d9; margin-bottom: 20px;">⚙️ Sobre o Sistema</h2>
        </div>
    """, unsafe_allow_html=True)
    
    st.markdown("""
        <div style="background-color: white; padding: 30px; border-radius: 10px; box-shadow: 0 2px 4px rgba(0,0,0,0.1);">
            <h3 style="color: #1f61d9; margin-bottom: 20px;">QD Apps - Sustentação</h3>
            <h4 style="color: #666; margin-bottom: 15px;">Versão 2.5</h4>
            <p style="color: #444; margin-bottom: 20px;">Sistema desenvolvido para auxiliar na gestão e controle das CHGs do time de Sustentação QD Apps.</p>
            <h4 style="color: #666; margin-bottom: 15px;">Funcionalidades Disponíveis:</h4>
            <ul style="list-style-type: none; padding-left: 0;">
                <li style="margin-bottom: 10px;">✨ Geração automática de relatórios de CHGs para o Keep</li>
                <li style="margin-bottom: 10px;">📅 Controle diário de CHGs agendadas</li>
                <li style="margin-bottom: 10px;">📱 Interface moderna e intuitiva</li>
            </ul>
            <div style="margin-top: 30px; padding-top: 20px; border-top: 1px solid #eee;">
                <p style="color: #666; font-size: 0.9em;">© 2024 Time de Sustentação QD Apps - Spread</p>
                <p style="color: #666; font-size: 0.9em; margin-top: 10px;">Desenvolvido por Mateus</p>
            </div>
        </div>
    """, unsafe_allow_html=True)
